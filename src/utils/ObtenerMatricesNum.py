from PIL import Image, ImageDraw, ImageFont
import openpyxl
import pandas as pd

def GenerarMatrices(ruta_imagen_entrada, ruta_salida_excel, ruta_salida_mapeo, ruta_salida_imagen_numerica, numero_inicial, tamañopixel,numeromaxpaisa):
    imagen_original = Image.open(ruta_imagen_entrada)
    ancho, alto = imagen_original.size

    tamano_pixel = tamañopixel

    libro_color = openpyxl.Workbook()
    hoja_color = libro_color.active

    color_a_numero = {}
    numero_a_color = {}
    numero_actual = numero_inicial  # Inicializa el número con el valor inicial dado

    # Crear una nueva imagen para la matriz numérica
    imagen_numerica = Image.new('RGB', (ancho, alto), (255, 255, 255))
    draw = ImageDraw.Draw(imagen_numerica)
    font = ImageFont.load_default()

    # pixelar_imagen reduce con división entera (ancho // tamano_pixel bloques) y
    # reescala de vuelta, así que el tamaño real de bloque es ancho / n_bloques (no
    # tamano_pixel). Se muestrea el centro de cada bloque real para que la matriz
    # corresponda 1:1 con el pixel art, sin filas/columnas duplicadas artificialmente.
    bloques_x = max(1, ancho // tamano_pixel)
    bloques_y = max(1, alto // tamano_pixel)
    bloque_ancho = ancho / bloques_x
    bloque_alto = alto / bloques_y

    for i in range(bloques_y):
        for j in range(bloques_x):
            muestra_x = int((j + 0.5) * bloque_ancho)
            muestra_y = int((i + 0.5) * bloque_alto)
            color_pixel = imagen_original.getpixel((muestra_x, muestra_y))

            # Asigna 0 si es transparente
            if color_pixel == (255, 255, 255, 0) or color_pixel == (0, 0, 0, 0):
                hoja_color.cell(row=i + 1, column=j + 1, value=0)
                draw.text((int(j * bloque_ancho), int(i * bloque_alto)), "0", fill=(0, 0, 0), font=font)
            else:
                # Si el color ya tiene un número asignado, lo reutiliza
                if color_pixel not in color_a_numero:
                    color_a_numero[color_pixel] = numero_actual
                    numero_a_color[numero_actual] = color_pixel
                    numero_actual = numeromaxpaisa+numero_actual+1
                # Escribe el número asignado en el Excel y en la imagen de salida
                hoja_color.cell(row=i + 1, column=j + 1, value=color_a_numero[color_pixel])
                draw.text((int(j * bloque_ancho), int(i * bloque_alto)), str(color_a_numero[color_pixel]), fill=(0, 0, 0), font=font)

    libro_color.save(ruta_salida_excel)

    # Crear el mapeo de número a color en un archivo Excel
    df_mapeo_color = pd.DataFrame(list(numero_a_color.items()), columns=['Número', 'Color (RGB)'])
    df_mapeo_color.to_excel(ruta_salida_mapeo, index=False)

    # Guardar la imagen de la matriz numérica
    imagen_numerica.save(ruta_salida_imagen_numerica)

    print("Procesamiento completado para:", ruta_imagen_entrada)

    return numero_actual - 1, numero_a_color





